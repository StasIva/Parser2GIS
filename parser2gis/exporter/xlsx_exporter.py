from __future__ import annotations

from typing import Any, Callable

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from parser2gis.exporter.base import BaseExporter

EXPORT_COLUMNS: list[str] = [
    "ID",
    "Название",
    "Город",
    "Адрес",
    "Телефоны",
    "Email",
    "Сайт",
    "Рубрика",
    "Время работы",
    "Координаты",
    "Источник ID",
]


class XlsxExporter(BaseExporter):
    def export(self, records: list[dict[str, Any]], file_path: str,
               on_progress: Callable[[int, int], None] | None = None) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Organizations"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, col_name in enumerate(EXPORT_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        for row_idx, record in enumerate(records, 2):
            ws.cell(row=row_idx, column=1, value=record.get("id", ""))
            ws.cell(row=row_idx, column=2, value=record.get("name", ""))
            ws.cell(row=row_idx, column=3, value=record.get("city", ""))
            ws.cell(row=row_idx, column=4, value=record.get("address", ""))
            phone_cell = ws.cell(row=row_idx, column=5, value=record.get("phones", ""))
            phone_cell.number_format = "@"
            ws.cell(row=row_idx, column=6, value=record.get("emails", ""))
            ws.cell(row=row_idx, column=7, value=record.get("website", ""))
            ws.cell(row=row_idx, column=8, value=record.get("rubric_name", ""))
            ws.cell(row=row_idx, column=9, value=record.get("work_hours", ""))
            coords = ""
            lat = record.get("lat")
            lon = record.get("lon")
            if lat is not None and lon is not None:
                coords = f"{lat}, {lon}"
            ws.cell(row=row_idx, column=10, value=coords)
            ws.cell(row=row_idx, column=11, value=record.get("source_id", ""))

            if on_progress:
                on_progress(row_idx - 1, len(records))

        for col_idx in range(1, len(EXPORT_COLUMNS) + 1):
            col_letter = get_column_letter(col_idx)
            max_length = len(EXPORT_COLUMNS[col_idx - 1])
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2, max_row=ws.max_row):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

        wb.save(file_path)
        return file_path